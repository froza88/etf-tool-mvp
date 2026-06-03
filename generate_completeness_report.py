#!/usr/bin/env python3
"""
ETF数据完整度总表生成器
- 读取 etf_standard_data.json (1490只ETF)
- 检查34个目标字段的完整度
- 输出Excel报告（2个工作表：主表 + 统计摘要）
- 输出HTML交互式报告（筛选、排序、图表）
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

# 项目根目录
PROJECT_ROOT = Path(__file__).parent
DATA_FILE = PROJECT_ROOT / "etf_standard_data.json"

# 25个目标字段（核心字段+Wind核心字段，字段名与etf_standard_data.json实际字段名一致）
TARGET_FIELDS = [
    # 基础信息
    'code', 'name', 'issuer', 'issuer_full', 'issuer_short',
    'scale', 'shares', 'issue_date', 'custodian', 'fund_manager',
    # 行情数据
    'change_pct', 'close', 'prev_close', 'change_rate',
    'volume', 'year_1_return', 'year_3_return',
    # 风险指标
    'max_drawdown', 'sharpe_ratio', 'annual_vol', 'category',
    # Wind数据
    'wind_基金管理人', 'wind_基金托管人', 'wind_基金成立日', 'wind_业绩比较基准'
]

# 确保openpyxl可用
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️ 缺少openpyxl库，正在安装...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"],
                   capture_output=True, check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


def load_etf_data():
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    etfs = data if isinstance(data, list) else data['etfs']
    return etfs


def is_field_missing(value, field=''):
    """判断字段值是否缺失（field参数用于特殊字段处理逻辑）"""
    # 通用缺失判断：None、空字符串、空列表/字典
    if value is None or value == '':
        return True
    if isinstance(value, (list, dict)) and len(value) == 0:
        return True
    
    # 特殊字段处理
    if field == 'volume':
        # volume=0 可能是停牌/新上市，判定为缺失
        return value == 0
    
    # 其他数值字段：0 可能是有效值，不判定为缺失
    return False


def check_field_completeness(etf):
    results = {}
    missing_fields = []
    for field in TARGET_FIELDS:
        value = etf.get(field)
        missing = is_field_missing(value, field)
        results[field] = {'value': value, 'is_missing': missing}
        if missing:
            missing_fields.append(field)
    rate = (len(TARGET_FIELDS) - len(missing_fields)) / len(TARGET_FIELDS)
    return results, missing_fields, rate


def generate_excel_report(etfs, output_file):
    print(f"📊 开始生成Excel报告（{len(etfs)}只ETF x {len(TARGET_FIELDS)}字段）...")
    wb = openpyxl.Workbook()

    # ===== Sheet 1: 主表 =====
    ws1 = wb.active
    ws1.title = "ETF数据完整度总表"
    _write_main_sheet(ws1, etfs)
    print(f"  ✓ Sheet1 'ETF数据完整度总表' 已生成（{ws1.max_row}行 x {ws1.max_column}列）")

    # ===== Sheet 2: 统计摘要 =====
    ws2 = wb.create_sheet("统计摘要")
    _write_stats_sheet(ws2, etfs)
    print(f"  ✓ Sheet2 '统计摘要' 已生成")

    wb.save(output_file)
    print(f"✅ Excel报告已保存: {output_file.name}")
    return output_file


def _write_main_sheet(ws, etfs):
    header_font = Font(bold=True, size=10)
    missing_fill = PatternFill(start_color="FFE7E6", end_color="FFE7E6", fill_type="solid")
    ok_fill = PatternFill(start_color="E7F6E7", end_color="E7F6E7", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ['ETF代码', 'ETF名称', '完整度%', '缺失字段数', '缺失字段列表'] + TARGET_FIELDS
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for row_idx, etf in enumerate(etfs, 2):
        results, missing_fields, rate = check_field_completeness(etf)
        ws.cell(row=row_idx, column=1, value=etf.get('code', ''))
        ws.cell(row=row_idx, column=2, value=etf.get('name', ''))
        ws.cell(row=row_idx, column=3, value=round(rate*100, 1))
        ws.cell(row=row_idx, column=4, value=len(missing_fields))
        ws.cell(row=row_idx, column=5, value=', '.join(missing_fields) if missing_fields else '无')
        for col_idx, field in enumerate(TARGET_FIELDS, 6):
            r = results.get(field, {'value': None, 'is_missing': True})
            val = r['value']
            display = str(val)[:50] if val is not None else '缺失'
            c = ws.cell(row=row_idx, column=col_idx, value=display)
            c.fill = missing_fill if r['is_missing'] else ok_fill

    # 自动列宽
    for col in range(1, len(headers)+1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, min(10, ws.max_row+1)):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    ws.freeze_panes = "F2"


def _write_stats_sheet(ws, etfs):
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    row = 1

    # ---- 区块1：各字段缺失统计 ----
    ws.cell(row=row, column=1, value="各字段缺失统计").font = Font(bold=True, size=13)
    ws.row_dimensions[row].height = 22
    row += 1

    for col, h in enumerate(['字段名', '缺失数量', '缺失比例', '有数据数量', '有数据比例'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    row += 1

    field_stats = {}
    for field in TARGET_FIELDS:
        missing_cnt = sum(1 for e in etfs if is_field_missing(e.get(field)))
        total = len(etfs)
        ok_cnt = total - missing_cnt
        field_stats[field] = {'missing': missing_cnt, 'ok': ok_cnt, 'missing_ratio': missing_cnt/total*100}

    for i, (field, stats) in enumerate(field_stats.items()):
        fill = alt_fill if i % 2 == 0 else white_fill
        ws.cell(row=row, column=1, value=field).alignment = left
        ws.cell(row=row, column=2, value=stats['missing']).alignment = center
        ws.cell(row=row, column=3, value=f"{stats['missing_ratio']:.1f}%").alignment = center
        ws.cell(row=row, column=4, value=stats['ok']).alignment = center
        ws.cell(row=row, column=5, value=f"{100-stats['missing_ratio']:.1f}%").alignment = center
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    row += 2

    # ---- 区块2：完整度分布 ----
    ws.cell(row=row, column=1, value="完整度分布").font = Font(bold=True, size=13)
    ws.row_dimensions[row].height = 22
    row += 1

    bins = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 95), (95, 101)]
    bin_labels = ['0-20%', '20-40%', '40-60%', '60-80%', '80-95%', '95-100%']
    bin_counts = {label: 0 for label in bin_labels}
    for etf in etfs:
        _, _, rate = check_field_completeness(etf)
        pct = rate * 100
        for (lo, hi), label in zip(bins, bin_labels):
            if lo <= pct < hi:
                bin_counts[label] += 1
                break

    for col, h in enumerate(['完整度区间', 'ETF数量', '占比'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center
    row += 1

    for i, label in enumerate(bin_labels):
        cnt = bin_counts[label]
        ratio = cnt / len(etfs) * 100
        fill = alt_fill if i % 2 == 0 else white_fill
        ws.cell(row=row, column=1, value=label).alignment = left
        ws.cell(row=row, column=2, value=cnt).alignment = center
        ws.cell(row=row, column=3, value=f"{ratio:.1f}%").alignment = center
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    row += 2

    # ---- 区块3：完整度最低Top10 ----
    ws.cell(row=row, column=1, value="完整度最低 Top 10").font = Font(bold=True, size=13)
    ws.row_dimensions[row].height = 22
    row += 1

    sorted_etfs = sorted(etfs, key=lambda e: check_field_completeness(e)[2])
    for col, h in enumerate(['排名', 'ETF代码', 'ETF名称', '完整度%', '缺失字段'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center
    row += 1

    for i, etf in enumerate(sorted_etfs[:10], 1):
        _, miss, rate = check_field_completeness(etf)
        fill = alt_fill if i % 2 == 0 else white_fill
        ws.cell(row=row, column=1, value=i).alignment = center; ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=etf.get('code','')).alignment = center; ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=3, value=etf.get('name','')).alignment = left; ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=4, value=f"{rate*100:.1f}%").alignment = center; ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=5, value=', '.join(miss) if miss else '无').alignment = left; ws.cell(row=row, column=5).fill = fill
        row += 1

    # 自动列宽
    for col in range(1, 6):
        max_len = 0
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)


def generate_html_report(etfs, output_file):
    """生成HTML交互式报告（单独文件）"""
    print(f"🌐 开始生成HTML交互式报告...")

    # 准备数据
    etf_rows = []
    for etf in etfs:
        results, missing_fields, rate = check_field_completeness(etf)
        etf_rows.append({
            'code': etf.get('code', ''),
            'name': etf.get('name', ''),
            'completeness': round(rate*100, 1),
            'missing_count': len(missing_fields),
            'missing_fields': missing_fields,
        })

    # 字段缺失统计
    field_stats = {}
    for field in TARGET_FIELDS:
        missing_cnt = sum(1 for e in etfs if is_field_missing(e.get(field)))
        field_stats[field] = {'missing': missing_cnt, 'ok': len(etfs)-missing_cnt, 'missing_ratio': missing_cnt/len(etfs)*100}

    # 完整度分布
    bins = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 95), (95, 101)]
    bin_labels = ['0-20%', '20-40%', '40-60%', '60-80%', '80-95%', '95-100%']
    bin_counts = {label: 0 for label in bin_labels}
    for r in etf_rows:
        pct = r['completeness']
        for (lo, hi), label in zip(bins, bin_labels):
            if lo <= pct < hi:
                bin_counts[label] += 1
                break

    # 写入HTML文件
    html_path = output_file.with_suffix('.html')
    _build_html(html_path, etf_rows, field_stats, bin_counts, bin_labels)
    print(f"✅ HTML报告已保存: {html_path.name}")
    return html_path


def _build_html(output_path, etf_rows, field_stats, bin_counts, bin_labels):
    import json
    etf_json = json.dumps(etf_rows, ensure_ascii=False)
    field_stats_json = json.dumps(field_stats, ensure_ascii=False)
    bin_data = json.dumps([bin_counts[l] for l in bin_labels], ensure_ascii=False)
    bin_labels_json = json.dumps(bin_labels, ensure_ascii=False)

    html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ETF数据完整度报告</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family: -apple-system, 'Segoe UI', sans-serif; background:#f5f7fa; color:#333; }
.header { background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); color:white; padding:24px 32px; }
.header h1 { font-size:24px; margin-bottom:4px; }
.header p { font-size:14px; opacity:0.85; }
.nav { display:flex; gap:8px; padding:12px 32px; background:white; border-bottom:1px solid #e5e7eb; }
.nav-btn { padding:8px 18px; border:1px solid #d1d5db; background:white; border-radius:6px; cursor:pointer; font-size:13px; }
.nav-btn.active { background:#3b82f6; color:white; border-color:#3b82f6; }
.section { padding:24px 32px; }
.card { background:white; border-radius:12px; padding:20px; margin-bottom:20px; box-shadow:0 1px 3px rgba(0,0,0,0.08); }
.card h2 { font-size:16px; margin-bottom:16px; color:#1e3a8a; }
.stats-grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(160px,1fr)); gap:12px; }
.stat-box { background:#f0f9ff; border-radius:8px; padding:14px; text-align:center; }
.stat-box .num { font-size:28px; font-weight:700; color:#1e3a8a; }
.stat-box .label { font-size:12px; color:#6b7280; margin-top:4px; }
table { width:100%; border-collapse:collapse; font-size:13px; }
th { background:#f3f4f6; padding:8px 10px; text-align:left; font-weight:600; border-bottom:2px solid #e5e7eb; position:sticky; top:0; cursor:pointer; }
td { padding:8px 10px; border-bottom:1px solid #f3f4f6; }
tr:hover { background:#f9fafb; }
.missing { color:#ef4444; font-weight:600; }
.ok { color:#10b981; }
.filter-bar { display:flex; gap:10px; margin-bottom:16px; flex-wrap:wrap; align-items:center; }
.filter-bar input, .filter-bar select { padding:7px 12px; border:1px solid #d1d5db; border-radius:6px; font-size:13px; }
.filter-bar label { font-size:13px; color:#6b7280; }
.badge-miss { background:#fef2f2; color:#ef4444; padding:2px 8px; border-radius:10px; font-size:11px; }
.badge-ok { background:#f0fdf4; color:#10b981; padding:2px 8px; border-radius:10px; font-size:11px; }
.chart-container { display:flex; gap:20px; flex-wrap:wrap; }
.chart-box { flex:1; min-width:300px; }
.chart-bar { display:flex; align-items:center; margin-bottom:6px; }
.chart-bar .label { width:80px; font-size:12px; text-align:right; padding-right:8px; }
.chart-bar .bar { height:18px; background:#3b82f6; border-radius:3px; min-width:2px; }
.chart-bar .num { font-size:11px; margin-left:6px; color:#6b7280; }
.pagination { display:flex; gap:6px; align-items:center; margin-top:12px; }
.pagination button { padding:5px 10px; border:1px solid #d1d5db; background:white; border-radius:4px; cursor:pointer; font-size:12px; }
.pagination button:disabled { opacity:0.4; cursor:default; }
</style>
</head>
<body>
<div class="header">
  <h1>📊 ETF数据完整度报告</h1>
  <p>共 <strong id="totalCount">0</strong> 只ETF | <strong id="avgCompleteness">0</strong>% 平均完整度 | 生成时间：<span id="genTime"></span></p>
</div>
<div class="nav">
  <button class="nav-btn active" onclick="showTab('overview')">总览</button>
  <button class="nav-btn" onclick="showTab('fields')">字段分析</button>
  <button class="nav-btn" onclick="showTab('etflist')">ETF列表</button>
</div>

<!-- Tab: 总览 -->
<div id="tab-overview" class="section">
  <div class="stats-grid" id="statsGrid"></div>
  <div class="card" style="margin-top:20px;">
    <h2>完整度分布</h2>
    <div class="chart-container">
      <div class="chart-box"><div id="distChart"></div></div>
      <div class="chart-box"><div id="fieldChart"></div></div>
    </div>
  </div>
</div>

<!-- Tab: 字段分析 -->
<div id="tab-fields" class="section" style="display:none;">
  <div class="card">
    <h2>各字段缺失统计（点击表头排序）</h2>
    <table>
      <thead><tr>
        <th onclick="sortFieldTable('field')">字段名 ↕</th>
        <th onclick="sortFieldTable('missing')">缺失数量 ↕</th>
        <th onclick="sortFieldTable('missingRatio')">缺失比例 ↕</th>
        <th onclick="sortFieldTable('ok')">有数据 ↕</th>
        <th>缺失比例条</th>
      </tr></thead>
      <tbody id="fieldTableBody"></tbody>
    </table>
  </div>
</div>

<!-- Tab: ETF列表 -->
<div id="tab-etflist" class="section" style="display:none;">
  <div class="filter-bar">
    <label>筛选：</label>
    <select id="filterCompleteness" onchange="renderETFList()">
      <option value="all">全部完整度</option>
      <option value="0-20">0-20%</option>
      <option value="20-40">20-40%</option>
      <option value="40-60">40-60%</option>
      <option value="60-80">60-80%</option>
      <option value="80-95">80-95%</option>
      <option value="95-100">95-100%</option>
    </select>
    <input type="text" id="filterCode" placeholder="搜ETF代码/名称" oninput="renderETFList()" style="width:160px;">
    <label>每页：</label>
    <select id="pageSize" onchange="renderETFList()">
      <option value="20">20</option>
      <option value="50">50</option>
      <option value="100">100</option>
    </select>
  </div>
  <div class="card">
    <table>
      <thead><tr>
        <th>代码</th><th>名称</th><th onclick="sortETFList('completeness')">完整度% ↕</th>
        <th>缺失数</th><th>缺失字段</th>
      </tr></thead>
      <tbody id="etfTableBody"></tbody>
    </table>
    <div class="pagination" id="pagination"></div>
  </div>
</div>

<script>
var ETF_DATA = """ + etf_json + """;
var FIELD_STATS = """ + field_stats_json + """;
var BIN_LABELS = """ + bin_labels_json + """;
var BIN_DATA = """ + bin_data + """;
var currentTab = 'overview';
var fieldSortKey = 'missing';
var fieldSortDir = -1;
var etfSortKey = 'completeness';
var etfSortDir = -1;
var etfPage = 1;

document.getElementById('genTime').textContent = new Date().toLocaleString('zh-CN');

function showTab(tab) {
  currentTab = tab;
  var btns = document.querySelectorAll('.nav-btn');
  for (var i=0; i<btns.length; i++) btns[i].classList.remove('active');
  event.target.classList.add('active');
  document.getElementById('tab-overview').style.display = tab=='overview'?'block':'none';
  document.getElementById('tab-fields').style.display = tab=='fields'?'block':'none';
  document.getElementById('tab-etflist').style.display = tab=='etflist'?'block':'none';
  if (tab=='overview' && !window._overviewRendered) { renderOverview(); window._overviewRendered=true; }
  if (tab=='fields' && !window._fieldsRendered) { renderFieldTable(); window._fieldsRendered=true; }
  if (tab=='etflist' && !window._etfRendered) { renderETFList(); window._etfRendered=true; }
}

function renderOverview() {
  var total = ETF_DATA.length;
  var avg = (ETF_DATA.reduce(function(s,e){return s+e.completeness},0)/total).toFixed(1);
  document.getElementById('totalCount').textContent = total;
  document.getElementById('avgCompleteness').textContent = avg;

  var statsGrid = document.getElementById('statsGrid');
  var stats = [
    ['ETF总数', total, '#1e3a8a'],
    ['平均完整度', avg+'%', '#059669'],
    ['完整度≥80%', ETF_DATA.filter(function(e){return e.completeness>=80}).length, '#7c3aed'],
    ['完整度<50%', ETF_DATA.filter(function(e){return e.completeness<50}).length, '#dc2626'],
  ];
  statsGrid.innerHTML = stats.map(function(s){return '<div class="stat-box"><div class="num" style="color:'+s[2]+';">'+s[1]+'</div><div class="label">'+s[0]+'</div></div>'}).join('');

  var distChart = document.getElementById('distChart');
  var maxCount = Math.max.apply(null, BIN_DATA);
  distChart.innerHTML = BIN_LABELS.map(function(l,i){
    var w = maxCount>0 ? Math.round(BIN_DATA[i]/maxCount*200) : 0;
    return '<div class="chart-bar"><div class="label">'+l+'</div><div class="bar" style="width:'+w+'px"></div><div class="num">'+BIN_DATA[i]+'只 ('+((BIN_DATA[i]/total)*100).toFixed(1)+'%)</div></div>';
  }).join('');

  var fieldArr = Object.entries(FIELD_STATS).sort(function(a,b){return b[1].missing - a[1].missing});
  var fieldChart = document.getElementById('fieldChart');
  fieldChart.innerHTML = '<h3 style="font-size:13px;margin-bottom:10px;color:#6b7280;">缺失最多的字段 Top 10</h3>' +
    fieldArr.slice(0,10).map(function(f){
      var w = f[1].missing_ratio>0 ? Math.round(f[1].missing_ratio/100*200) : 0;
      return '<div class="chart-bar"><div class="label" style="font-size:11px;">'+f[0]+'</div><div class="bar" style="width:'+w+'px;background:#ef4444"></div><div class="num">'+f[1].missing+'只 ('+f[1].missing_ratio.toFixed(1)+'%)</div></div>';
    }).join('');
}

function renderFieldTable() {
  var tbody = document.getElementById('fieldTableBody');
  var arr = Object.entries(FIELD_STATS).sort(function(a,b){
    var vA = a[1][fieldSortKey], vB = b[1][fieldSortKey];
    return fieldSortDir * (vA>vB?1:vA<vB?-1:0);
  });
  tbody.innerHTML = arr.map(function(f){return '<tr><td>'+f[0]+'</td><td class="missing">'+f[1].missing+'</td><td>'+f[1].missing_ratio.toFixed(1)+'%</td><td class="ok">'+f[1].ok+'</td><td><div style="background:#fee2e2;height:8px;width:'+f[1].missing_ratio+'%;border-radius:4px;"></div></td></tr>'}).join('');
}

function sortFieldTable(key) {
  if (fieldSortKey===key) fieldSortDir *= -1; else { fieldSortKey=key; fieldSortDir=-1; }
  renderFieldTable();
}

function renderETFList() {
  var filterRange = document.getElementById('filterCompleteness').value;
  var filterText = document.getElementById('filterCode').value.toLowerCase();
  var pageSize = parseInt(document.getElementById('pageSize').value);

  var filtered = ETF_DATA.filter(function(e){
    if (filterRange!='all') {
      var parts = filterRange.split('-'); var lo=parseInt(parts[0]); var hi=parseInt(parts[1]);
      if (e.completeness < lo || e.completeness >= hi) return false;
    }
    if (filterText && !(e.code.toLowerCase().indexOf(filterText)>=0 || e.name.toLowerCase().indexOf(filterText)>=0)) return false;
    return true;
  });

  filtered.sort(function(a,b){
    var vA = a[etfSortKey], vB = b[etfSortKey];
    return etfSortDir * (vA>vB?1:vA<vB?-1:0);
  });

  var totalPages = Math.ceil(filtered.length / pageSize);
  etfPage = Math.min(etfPage, totalPages || 1);
  var start = (etfPage-1)*pageSize;
  var pageData = filtered.slice(start, start+pageSize);

  var tbody = document.getElementById('etfTableBody');
  tbody.innerHTML = pageData.map(function(e){return '<tr><td><strong>'+e.code+'</strong></td><td>'+e.name+'</td><td class="'+(e.completeness>=80?'ok':'missing')+'">'+e.completeness+'%</td><td><span class="badge-miss">'+e.missing_count+'</span></td><td style="font-size:11px;color:#6b7280;">'+(e.missing_fields.length>0?e.missing_fields.join(', '):'<span class="ok">完整</span>')+'</td></tr>'}).join('');

  var pag = document.getElementById('pagination');
  pag.innerHTML = '';
  if (etfPage>1) pag.innerHTML += '<button onclick="etfPage--;renderETFList()">上一页</button>';
  pag.innerHTML += '<span style="font-size:12px;color:#6b7280;">第'+etfPage+'/'+totalPages+'页，共'+filtered.length+'只</span>';
  if (etfPage<totalPages) pag.innerHTML += '<button onclick="etfPage++;renderETFList()">下一页</button>';
}

function sortETFList(key) {
  if (etfSortKey===key) etfSortDir *= -1; else { etfSortKey=key; etfSortDir=-1; }
  etfPage=1;
  renderETFList();
}

// 初始化
renderOverview();
window._overviewRendered = true;
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ HTML报告已保存: {output_path.name}")


def main():
    print("=" * 60)
    print("ETF数据完整度总表生成器")
    print("=" * 60)

    print(f"\n📂 加载ETF数据: {DATA_FILE}")
    etfs = load_etf_data()
    print(f"   加载成功: {len(etfs)} 只ETF")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 生成Excel（含统计摘要）
    excel_file = PROJECT_ROOT / f"etf_completeness_report_{timestamp}.xlsx"
    generate_excel_report(etfs, excel_file)

    # 生成HTML交互式报告
    html_file = PROJECT_ROOT / f"etf_completeness_report_{timestamp}.html"
    generate_html_report(etfs, html_file)

    print(f"\n✅ 全部完成！")
    print(f"   Excel报告: {excel_file.name}")
    print(f"   HTML报告: {html_file.name}（用浏览器打开）")


if __name__ == "__main__":
    main()
